import string
import pandas as pd
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side, PatternFill
import numpy as np


def AddTableBordersAndGrid(worksheet, df):
    #Find full range of table

    cols = GetDFColumns(df)
    last_column = cols[1]
    rows = df.shape[0] + 1
    rng = 'A1:'+last_column+str(rows)

    #Add Cell Border
    rows = worksheet[rng]
    for row in rows:
        for c in row:
            if c.column_letter == last_column:
                c.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='medium'))
            elif c.column_letter == 'A':
                c.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='medium'), right=Side(style='thin'))
            else:
                c.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    #Add top/bottom border
    for c in rows[0]:         
        if c.column_letter == 'A':
            c.border = Border(top=Side(style='medium'), bottom=Side(style='thin'), left=Side(style='medium'), right=Side(style='thin'))
        elif c.column_letter == last_column:
            c.border = Border(top=Side(style='medium'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='medium'))
        else:
            c.border = Border(top=Side(style='medium'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
    for c in rows[-1]: 
        if c.column_letter == 'A':
            c.border = Border(top=Side(style='thin'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='thin'))
        elif c.column_letter == last_column:
            c.border = Border(top=Side(style='thin'), bottom=Side(style='medium'), left=Side(style='thin'), right=Side(style='medium'))
        else:
            c.border = Border(top=Side(style='thin'), bottom=Side(style='medium'), left=Side(style='thin'), right=Side(style='thin'))
    return worksheet


def AutoFitColumnsAndHeaders( worksheet, df, header_row ):
    #Estimate the height of the header row

    #Adjust the rows not at the column header row
    for column in worksheet.iter_cols():
        col = column[0].col_idx
        letter = column[0].column_letter
        max_cell_size = 30
        min_cell_size = 8.43
        for row in worksheet.iter_rows(min_col = col , max_col = col , min_row = 1):                                       
            cell_size = len(str(row[0].internal_value))
            row[0].alignment = Alignment(horizontal='center', wrapText=True)
            if cell_size > min_cell_size and cell_size < max_cell_size:
                min_cell_size = cell_size

        worksheet.column_dimensions[letter].width = min_cell_size + 0.3           

    return worksheet

def AddRedGreenColorGradient(worksheet, header, df, hi, lo):
    d = dict(zip(range(25), list(string.ascii_uppercase)[1:]))

    excel_header = str(d[df.columns.get_loc(header) - 1])
    len_df = df.shape[0] + 1
    rng = excel_header + '2:' + excel_header + str(len_df)

    worksheet.conditional_formatting.add(rng, ColorScaleRule(start_type ='num', start_value = lo, start_color = '0000FF00',\
                                                                mid_type = 'num', mid_value = np.average([lo, hi]), mid_color = '00FFFF00',\
                                                                end_type = 'num', end_value = hi, end_color = '00FF0000'))

    return worksheet

def ShadeColumn(worksheet, column, color):
    color_idx = GetColorIndex(color)
    fill = PatternFill(start_color=color_idx, end_color=color_idx, fill_type='solid')

    for col in worksheet.iter_cols():
        for cell in col:
            cell.fill = fill

    return worksheet

def ShadeRow(worksheet, row_num, color):
    color_idx = GetColorIndex(color)
    fill = PatternFill(start_color=color_idx, end_color=color_idx, fill_type='solid')

    for row in worksheet.iter_rows(min_row = row_num, max_row = row_num):
        for cell in row:
            cell.fill = fill

    return worksheet

def GetColorIndex(color):
    COLOR_INDEX = (
        '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF', #0-4
        '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF', #5-9
        '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF', #10-14
        '0000FFFF', '00800000', '00008000', '00000080', '00808000', #15-19
        '00800080', '00008080', '00C0C0C0', '00808080', '009999FF', #20-24
        '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080', #25-29
        '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00', #30-34
        '0000FFFF', '00800080', '00800000', '00008080', '000000FF', #35-39
        '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF', #40-44
        '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC', #45-49
        '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699', #50-54
        '00969696', '00003366', '00339966', '00003300', '00333300', #55-59
        '00993300', '00993366', '00333399', '00333333', 'System Foreground', 'System Background' #60-64
    )

    color_dict = {}
    color_dict['BLACK'] = COLOR_INDEX[0]
    color_dict['WHITE'] = COLOR_INDEX[1]
    color_dict['RED'] = COLOR_INDEX[2]
    color_dict['DARKRED'] = COLOR_INDEX[8]
    color_dict['BLUE'] = COLOR_INDEX[4]
    color_dict['DARKBLUE'] = COLOR_INDEX[12]
    color_dict['GREEN'] = COLOR_INDEX[3]
    color_dict['DARKGREEN'] = COLOR_INDEX[9]
    color_dict['YELLOW'] = COLOR_INDEX[5]
    color_dict['DARKYELLOW'] = COLOR_INDEX[19]
    color_dict['GREY'] = COLOR_INDEX[23]

    try:
        color = color_dict[str(color).upper()]
    except:
        color = ''
    return color

def AddRowFilter(worksheet, filterrownumber):

    return worksheet

def AddImage(worksheet, image, anchor_pos):
    img = openpyxl.drawing.image.Image(image)
    img.anchor = anchor_pos

    worksheet.add_image(img)
    return worksheet

def GetDFColumns(df, rng = ''):
    
    d = GetExcelColumnLetters(df, len(df.columns))
    last_col = str(d[df.columns.get_loc(df.columns[-1])])
    first_col = str(d[df.columns.get_loc(df.columns[0])])

    return first_col, last_col, d

def HideColumns(worksheet, min_col, max_col):
    #Hide all columns in col_array
    for col in worksheet.iter_cols(min_col = min_col, max_col = max_col):
        worksheet.column_dimensions[col[0].column_letter].hidden = True
    return worksheet

def GetExcelColumnLetters(df, rng):
    import itertools

    if not rng:
        rng = len(df.columns)
        
    key = range(rng)
    values = list(itertools.chain(string.ascii_uppercase, (''.join(pair) for pair in itertools.product(string.ascii_uppercase, repeat=2))))
    d = dict(zip(key, values))

    return d